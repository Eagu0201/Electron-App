import sys, json, csv, shutil, os, openpyxl
from google.cloud import automl_v1beta1 as automl
from google.cloud.automl_v1beta1.proto import service_pb2
from google.oauth2 import service_account
from flask import Flask
from openpyxl import Workbook
from openpyxl import load_workbook

datos = sys.argv[1]

print(datos)

wb = load_workbook(filename = '../excelOutputs/output-test.xlsx')
print(wb.sheetnames)
ws = wb['UE']

ws['C5'].value = datos

wb.save('../excelOutputs/output-test.xlsx') 

os.environ["GOOGLE_APPLICATION_CREDENTIALS"]='../stunning-hue-255412-12337a72db24.json'

def get_prediction(content, project_id, model_id):
  compute_region = 'us-central1'
  automl_client = automl.AutoMlClient()
  prediction_client = automl.PredictionServiceClient()
  name = 'projects/{}/locations/us-central1/models/{}'.format(project_id, model_id)
  model_full_id = automl_client.model_path(project_id,compute_region ,model_id)
  payload = {'text_snippet': {'content': content, 'mime_type': 'text/plain'}}
  params = {}
  request = prediction_client.predict(name, payload, params)
  return request  # waits till request is returned

output = (get_prediction(datos, "stunning-hue-255412",  "TCN6348187867260561980"))
print(output)
sys.stdout.flush()