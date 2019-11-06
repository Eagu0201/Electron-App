import sys, csv, shutil, os, json, openpyxl, re
from flask import Flask
from openpyxl import Workbook
from openpyxl import load_workbook

scores = json.loads(sys.argv[1])

wb = load_workbook(filename = './excelOutputs/output-test.xlsx')
ws = wb['UE']

ws['I5'].value = scores['AnalisisFuncional']
ws['J5'].value = scores['Arquitectura']
ws['K5'].value = scores['Programacion']
ws['L5'].value = scores['Testing']
ws['M5'].value = scores['Implementacion']
ws['N5'].value = scores['Gestion']
ws['O5'].value = scores['UX']

wb.save('./excelOutputs/output-test.xlsx') 

sys.stdout.flush()